!pip install pandas openpyxl numpy odfpy -q

# ANALISIS DE MICOTOXINAS

# Este script analiza datos de micotoxinas de un archivo Excel, clasifica las
# muestras como contaminadas o no contaminadas, y genera un informe detallado
# con tres tablas: resumen por año (matriz), resumen por mes (matriz),
# y tabla comparativa que muestra la contaminación de cada mes de cada año por micotoxina.
# Además, cuenta las muestras que tienen más de una micotoxina detectada.


# IMPORTAR LIBRERIAS


import pandas as pd
import numpy as np
from google.colab import files
import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

print("Librerias importadas correctamente")


# SUBIMOS ARCHIVO DESDE EL ORDENADOR


print("\nARCHIVO EXCEL:")
print("(Haz clic en 'Elegir archivo' y selecciona el archivo)")

uploaded = files.upload()

filename = list(uploaded.keys())[0]
print(f"\nArchivo subido: {filename}")
print(f"Tamaño: {len(uploaded[filename])} bytes")

# Leer el archivo Excel
df = pd.read_excel(io.BytesIO(uploaded[filename]))

print("\nVISTA PREVIA DE LOS DATOS:")
print(df.head())
print(f"\nTotal de filas: {len(df)}")
print(f"Total de columnas: {len(df.columns)}")

# Mostrar valores únicos de la columna de meses para depuración
print("\nVALORES ÚNICOS EN LA COLUMNA DE MES:")
meses_originales = df.iloc[:, 1].dropna().unique()
print(meses_originales)



# FUNCION PARA NORMALIZAR NOMBRES DE MESES


def normalizar_mes(mes):
    if pd.isna(mes):
        return None

    mes_str = str(mes).strip().upper()

    mapeo_meses = {
        'ENERO': 'ENERO', 'enero': 'ENERO', 'ENE': 'ENERO', 'ENERO.': 'ENERO',
        'FEBRERO': 'FEBRERO', 'febrero': 'FEBRERO', 'FEB': 'FEBRERO', 'FEBRERO.': 'FEBRERO',
        'MARZO': 'MARZO', 'marzo': 'MARZO', 'MAR': 'MARZO', 'MARZO.': 'MARZO',
        'ABRIL': 'ABRIL', 'abril': 'ABRIL', 'ABR': 'ABRIL', 'ABRIL.': 'ABRIL',
        'MAYO': 'MAYO', 'mayo': 'MAYO', 'MAY': 'MAYO', 'MAYO.': 'MAYO',
        'JUNIO': 'JUNIO', 'junio': 'JUNIO', 'JUN': 'JUNIO', 'JUNIO.': 'JUNIO',
        'JULIO': 'JULIO', 'julio': 'JULIO', 'JUL': 'JULIO', 'JULIO.': 'JULIO',
        'AGOSTO': 'AGOSTO', 'agosto': 'AGOSTO', 'AGO': 'AGOSTO', 'AGOSTO.': 'AGOSTO',
        'SEPTIEMBRE': 'SEPTIEMBRE', 'septiembre': 'SEPTIEMBRE', 'SEP': 'SEPTIEMBRE', 'SET': 'SEPTIEMBRE',
        'OCTUBRE': 'OCTUBRE', 'octubre': 'OCTUBRE', 'OCT': 'OCTUBRE', 'OCTUBRE.': 'OCTUBRE',
        'NOVIEMBRE': 'NOVIEMBRE', 'noviembre': 'NOVIEMBRE', 'NOV': 'NOVIEMBRE', 'NOVIEMBRE.': 'NOVIEMBRE',
        'DICIEMBRE': 'DICIEMBRE', 'diciembre': 'DICIEMBRE', 'DIC': 'DICIEMBRE', 'DICIEMBRE.': 'DICIEMBRE'
    }

    for clave, valor in mapeo_meses.items():
        if clave in mes_str or mes_str == clave:
            return valor

    return mes_str

# FUNCION PARA DETECTAR LA CONTAMINACIÓN DE LAS MICOTOXINAS

def esta_contaminado(valor):
    """
    Determina si una muestra esta contaminada:
    - Contaminado: valor numerico (sin <), o valor que empieza con '>'
    - No contaminado: vacio, empieza con <, o texto no numerico (excepto >)
    """
    if pd.isna(valor) or valor == "":
        return 0

    valor_str = str(valor).strip()

    if valor_str.startswith('<'):
        return 0

    if valor_str.startswith('>'):
        return 1

    try:
        valor_num = float(valor_str.replace(',', '.'))
        return 1
    except:
        return 0



# FUNCIÓN PARA CONTAR MICOTOXINAS POR MUESTRA


def contar_micotoxinas_por_muestra(row, columnas_toxinas):
    """
    Cuenta cuántas micotoxinas están presentes en una muestra
    """
    contador = 0
    for toxina in columnas_toxinas:
        if esta_contaminado(row[toxina]) == 1:
            contador += 1
    return contador


# PROCESAR DATOS

print("\n DATOS")

df.columns = [str(col).strip().upper() for col in df.columns]

mes_col_name = df.columns[1]
print(f"Columna de mes: {mes_col_name}")

df['MES_NORM'] = df.iloc[:, 1].apply(normalizar_mes)
df['MES'] = df['MES_NORM']
df.drop('MES_NORM', axis=1, inplace=True)

posibles_toxinas = ['AFB1', 'FUM', 'DON', 'ZEN', 'T-2', 'OTA']
columnas_toxinas = []

for col in df.columns:
    for toxina in posibles_toxinas:
        if toxina in col.upper():
            columnas_toxinas.append(col)
            break

print(f"Columnas de toxinas detectadas: {columnas_toxinas}")

if 'AÑO' not in df.columns:
    for col in df.columns:
        if 'AÑO' in col.upper() or 'YEAR' in col.upper():
            df.rename(columns={col: 'AÑO'}, inplace=True)
            break

años = sorted(df['AÑO'].unique())
print(f"Años encontrados: {años}")

todos_los_meses = ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO',
                   'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']

meses_con_datos = []
for mes in todos_los_meses:
    if mes in df['MES'].unique():
        meses_con_datos.append(mes)

print(f"Meses con datos: {meses_con_datos}")

# CONTAR MUESTRAS CON MÚLTIPLES MICOTOXINAS

print("\nMUESTRAS CON MÚLTIPLES MICOTOXINAS")

# Añadir columna con el número de micotoxinas por muestra
df['NUM_MICOTOXINAS'] = df.apply(lambda row: contar_micotoxinas_por_muestra(row, columnas_toxinas), axis=1)

# Contar muestras con más de 1 micotoxina
muestras_multitoxina = df[df['NUM_MICOTOXINAS'] > 1]
num_muestras_multitoxina = len(muestras_multitoxina)
total_muestras = len(df)

print(f"\nANÁLISIS DE MÚLTIPLES MICOTOXINAS")
print(f"Total de muestras analizadas: {total_muestras}")
print(f"Muestras con más de 1 micotoxina: {num_muestras_multitoxina}")
print(f"Porcentaje: {round(num_muestras_multitoxina / total_muestras * 100, 2)}%")

# Distribución por número de micotoxinas
print("\nDistribución por número de micotoxinas:")
distribucion = df['NUM_MICOTOXINAS'].value_counts().sort_index()
for n_micotoxinas, count in distribucion.items():
    porcentaje = round(count / total_muestras * 100, 2)
    if n_micotoxinas == 0:
        print(f"   Sin micotoxinas: {count} muestras ({porcentaje}%)")
    elif n_micotoxinas == 1:
        print(f"   Con 1 micotoxina: {count} muestras ({porcentaje}%)")
    else:
        print(f"   Con {n_micotoxinas} micotoxinas: {count} muestras ({porcentaje}%)")

# Análisis por año
print("\nAnálisis por año:")
for año in años:
    df_año = df[df['AÑO'] == año]
    total_año = len(df_año)
    multitoxina_año = len(df_año[df_año['NUM_MICOTOXINAS'] > 1])
    porcentaje_año = round(multitoxina_año / total_año * 100, 2) if total_año > 0 else 0
    print(f"   {año}: {multitoxina_año} muestras con múltiples micotoxinas / {total_año} total ({porcentaje_año}%)")

# Análisis por mes
print("\nAnálisis por mes:")
for mes in todos_los_meses:
    df_mes = df[df['MES'] == mes]
    total_mes = len(df_mes)
    if total_mes > 0:
        multitoxina_mes = len(df_mes[df_mes['NUM_MICOTOXINAS'] > 1])
        porcentaje_mes = round(multitoxina_mes / total_mes * 100, 2)
        print(f"   {mes}: {multitoxina_mes} muestras con múltiples micotoxinas / {total_mes} total ({porcentaje_mes}%)")

# Combinaciones más frecuentes
print("\nCombinaciones de micotoxinas más frecuentes:")
combinaciones = []
for idx, row in df[df['NUM_MICOTOXINAS'] > 1].iterrows():
    toxinas_presentes = [toxina for toxina in columnas_toxinas if esta_contaminado(row[toxina]) == 1]
    if len(toxinas_presentes) > 1:
        combinacion = ', '.join(sorted(toxinas_presentes))
        combinaciones.append(combinacion)

from collections import Counter
contador_combinaciones = Counter(combinaciones)
for combinacion, count in contador_combinaciones.most_common(10):
    print(f"   {combinacion}: {count} muestras")



# TABLA 1: RESUMEN POR AÑO (MATRIZ)

print("\nTABLA 1: RESUMEN POR AÑO (MATRIZ)")

resultados_anuales = []

for año in años:
    df_año = df[df['AÑO'] == año]
    total_muestras_año = len(df_año)

    for toxina in columnas_toxinas:
        contaminados = df_año[toxina].apply(esta_contaminado)
        num_contaminados = int(contaminados.sum())

        resultados_anuales.append({
            'Año': año,
            'Micotoxina': toxina,
            'Total Muestras': total_muestras_año,
            'Contaminados': num_contaminados,
            'No Contaminados': total_muestras_año - num_contaminados,
            '% Contaminacion': round(num_contaminados / total_muestras_año * 100, 2) if total_muestras_año > 0 else 0
        })

df_anual = pd.DataFrame(resultados_anuales)

# TABLA 2: RESUMEN POR MES (MATRIZ) - TODOS LOS MESES

print("TABLA 2: RESUMEN POR MES (MATRIZ)")

resultados_mensuales = []

for mes in todos_los_meses:
    df_mes = df[df['MES'] == mes]
    total_muestras_mes = len(df_mes)

    for toxina in columnas_toxinas:
        if total_muestras_mes > 0:
            contaminados = df_mes[toxina].apply(esta_contaminado)
            num_contaminados = int(contaminados.sum())
            porcentaje = round(num_contaminados / total_muestras_mes * 100, 2)
        else:
            num_contaminados = 0
            porcentaje = 0.0

        resultados_mensuales.append({
            'Mes': mes,
            'Micotoxina': toxina,
            'Total Muestras': total_muestras_mes,
            'Contaminados': num_contaminados,
            '% Contaminacion': porcentaje
        })

df_mensual = pd.DataFrame(resultados_mensuales)

# TABLA 3: DETALLE POR MES Y AÑO (PARA CONSOLA)

print("TABLA 3: DETALLE POR MES Y AÑO")

resultados_detalle = []

for año in años:
    for mes in todos_los_meses:
        df_mes_año = df[(df['AÑO'] == año) & (df['MES'] == mes)]
        total_muestras = len(df_mes_año)

        if total_muestras > 0:
            for toxina in columnas_toxinas:
                contaminados = df_mes_año[toxina].apply(esta_contaminado)
                num_contaminados = int(contaminados.sum())

                resultados_detalle.append({
                    'Año': año,
                    'Mes': mes,
                    'Micotoxina': toxina,
                    'Total Muestras': total_muestras,
                    'Contaminados': num_contaminados,
                    'No Contaminados': total_muestras - num_contaminados,
                    '% Contaminacion': round(num_contaminados / total_muestras * 100, 2)
                })
        else:
            for toxina in columnas_toxinas:
                resultados_detalle.append({
                    'Año': año,
                    'Mes': mes,
                    'Micotoxina': toxina,
                    'Total Muestras': 0,
                    'Contaminados': 0,
                    'No Contaminados': 0,
                    '% Contaminacion': 0.0
                })

df_detalle = pd.DataFrame(resultados_detalle)

# TABLA 4: COMPARACIÓN POR MES Y AÑO

# Esta tabla muestra para cada micotoxina, la contaminación de cada mes de cada año
# Formato: Filas = Meses, Columnas = Años (con dos subcolumnas: Contaminados y %)


print("TABLA 4: COMPARACIÓN POR MES Y AÑO")

# Crear diccionario para almacenar datos por micotoxina
datos_por_toxina = {}

for toxina in columnas_toxinas:
    # Crear estructura para esta toxina
    datos_toxina = {}
    for año in años:
        datos_toxina[año] = {}
        for mes in todos_los_meses:
            # Buscar datos
            filtro = (df_detalle['Año'] == año) & (df_detalle['Mes'] == mes) & (df_detalle['Micotoxina'] == toxina)
            datos = df_detalle[filtro]
            if len(datos) > 0:
                datos_toxina[año][mes] = {
                    'contaminados': int(datos['Contaminados'].iloc[0]),
                    'porcentaje': datos['% Contaminacion'].iloc[0]
                }
            else:
                datos_toxina[año][mes] = {
                    'contaminados': 0,
                    'porcentaje': 0.0
                }
    datos_por_toxina[toxina] = datos_toxina


# CREAR EXCEL CON LAS CUATRO TABLAS

print("\nGENERANDO ARCHIVO EXCEL")

output_filename = f"resumen_{filename.replace('.xlsx', '')}_micotoxinas_completo.xlsx"

wb = Workbook()


# ESTILOS

titulo_font = Font(bold=True, size=14, color="FFFFFF")
titulo_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

seccion_font = Font(bold=True, size=12, color="FFFFFF")
seccion_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="5490C1", end_color="5490C1", fill_type="solid")

center_alignment = Alignment(horizontal="center", vertical="center")
left_alignment = Alignment(horizontal="left", vertical="center")


# HOJA 1: RESUMEN POR AÑO (MATRIZ)

ws1 = wb.active
ws1.title = "Resumen_por_año"

fila = 1

ws1.merge_cells(f'A{fila}:F{fila}')
titulo = ws1.cell(row=fila, column=1, value="RESUMEN POR AÑO (MATRIZ)")
titulo.font = titulo_font
titulo.fill = titulo_fill
titulo.alignment = center_alignment
fila += 2

ws1.cell(row=fila, column=1, value="Micotoxina").font = header_font
ws1.cell(row=fila, column=1).fill = header_fill
ws1.cell(row=fila, column=1).alignment = center_alignment
ws1.cell(row=fila, column=2, value="Tipo").font = header_font
ws1.cell(row=fila, column=2).fill = header_fill
ws1.cell(row=fila, column=2).alignment = center_alignment

col = 3
for año in años:
    ws1.cell(row=fila, column=col, value=f"Cont {año}").font = header_font
    ws1.cell(row=fila, column=col).fill = header_fill
    ws1.cell(row=fila, column=col).alignment = center_alignment
    ws1.cell(row=fila, column=col+1, value=f"% {año}").font = header_font
    ws1.cell(row=fila, column=col+1).fill = header_fill
    ws1.cell(row=fila, column=col+1).alignment = center_alignment
    col += 2
fila += 1

toxinas = df_anual['Micotoxina'].unique()

for toxina in toxinas:
    ws1.cell(row=fila, column=1, value=toxina).alignment = left_alignment
    ws1.cell(row=fila, column=2, value="Contaminados").alignment = left_alignment

    col = 3
    for año in años:
        valor = df_anual[(df_anual['Año'] == año) & (df_anual['Micotoxina'] == toxina)]['Contaminados'].values
        porcentaje = df_anual[(df_anual['Año'] == año) & (df_anual['Micotoxina'] == toxina)]['% Contaminacion'].values
        if len(valor) > 0:
            ws1.cell(row=fila, column=col, value=int(valor[0])).alignment = center_alignment
            ws1.cell(row=fila, column=col+1, value=porcentaje[0]).alignment = center_alignment
        col += 2
    fila += 1

    ws1.cell(row=fila, column=2, value="No Contaminados").alignment = left_alignment
    col = 3
    for año in años:
        total = df_anual[(df_anual['Año'] == año) & (df_anual['Micotoxina'] == toxina)]['Total Muestras'].values
        contaminados = df_anual[(df_anual['Año'] == año) & (df_anual['Micotoxina'] == toxina)]['Contaminados'].values
        if len(total) > 0 and len(contaminados) > 0:
            no_cont = int(total[0] - contaminados[0])
            ws1.cell(row=fila, column=col, value=no_cont).alignment = center_alignment
        col += 2
    fila += 2

ws1.column_dimensions['A'].width = 25
ws1.column_dimensions['B'].width = 20
for i, año in enumerate(años):
    ws1.column_dimensions[chr(67 + i*2)].width = 12
    ws1.column_dimensions[chr(68 + i*2)].width = 10



# HOJA 2: RESUMEN POR MES - TODOS LOS MESES


ws2 = wb.create_sheet("Resumen_por_Mes")

fila = 1

ws2.merge_cells(f'A{fila}:F{fila}')
titulo = ws2.cell(row=fila, column=1, value="RESUMEN POR MES")
titulo.font = titulo_font
titulo.fill = titulo_fill
titulo.alignment = center_alignment
fila += 2

ws2.cell(row=fila, column=1, value="Micotoxina").font = header_font
ws2.cell(row=fila, column=1).fill = header_fill
ws2.cell(row=fila, column=1).alignment = center_alignment
ws2.cell(row=fila, column=2, value="Tipo").font = header_font
ws2.cell(row=fila, column=2).fill = header_fill
ws2.cell(row=fila, column=2).alignment = center_alignment

col = 3
for mes in todos_los_meses:
    ws2.cell(row=fila, column=col, value=f"Cont {mes[:4]}").font = header_font
    ws2.cell(row=fila, column=col).fill = header_fill
    ws2.cell(row=fila, column=col).alignment = center_alignment
    ws2.cell(row=fila, column=col+1, value=f"% {mes[:4]}").font = header_font
    ws2.cell(row=fila, column=col+1).fill = header_fill
    ws2.cell(row=fila, column=col+1).alignment = center_alignment
    col += 2
fila += 1

for toxina in toxinas:
    ws2.cell(row=fila, column=1, value=toxina).alignment = left_alignment
    ws2.cell(row=fila, column=2, value="Contaminados").alignment = left_alignment

    col = 3
    for mes in todos_los_meses:
        valor = df_mensual[(df_mensual['Mes'] == mes) & (df_mensual['Micotoxina'] == toxina)]['Contaminados'].values
        porcentaje = df_mensual[(df_mensual['Mes'] == mes) & (df_mensual['Micotoxina'] == toxina)]['% Contaminacion'].values
        if len(valor) > 0 and valor[0] > 0:
            ws2.cell(row=fila, column=col, value=int(valor[0])).alignment = center_alignment
            ws2.cell(row=fila, column=col+1, value=porcentaje[0]).alignment = center_alignment
        else:
            ws2.cell(row=fila, column=col, value="0").alignment = center_alignment
            ws2.cell(row=fila, column=col+1, value="0.0").alignment = center_alignment
        col += 2
    fila += 1

    ws2.cell(row=fila, column=2, value="No Contaminados").alignment = left_alignment
    col = 3
    for mes in todos_los_meses:
        total = df_mensual[(df_mensual['Mes'] == mes) & (df_mensual['Micotoxina'] == toxina)]['Total Muestras'].values
        contaminados = df_mensual[(df_mensual['Mes'] == mes) & (df_mensual['Micotoxina'] == toxina)]['Contaminados'].values
        if len(total) > 0 and len(contaminados) > 0:
            no_cont = int(total[0] - contaminados[0])
            ws2.cell(row=fila, column=col, value=no_cont).alignment = center_alignment
        else:
            ws2.cell(row=fila, column=col, value="0").alignment = center_alignment
        col += 2
    fila += 2

ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 20
for i in range(len(todos_los_meses)):
    ws2.column_dimensions[chr(67 + i*2)].width = 12
    ws2.column_dimensions[chr(68 + i*2)].width = 10


# HOJA 3: COMPARACIÓN POR MES Y AÑO (POR MICOTOXINA)

# Esta hoja muestra para cada micotoxina una matriz con meses en filas
# y años en columnas (cada año con Contaminados y %)


ws3 = wb.create_sheet("Comparacion_Mes_Ano")

# Crear una hoja por cada micotoxina
for idx, toxina in enumerate(columnas_toxinas):
    # Sanitize the sheet title to remove invalid characters like '/' and ensure max length
    # Replace '/' with '-' and truncate to 31 characters if necessary.
    safe_toxina_name = toxina.replace('/', '-').replace('*', '').replace('?', '').replace(':', '').replace('[', '').replace(']', '')
    sheet_title = f"Comp_{safe_toxina_name[:20]}"
    if idx == 0:
        ws = ws3
    else:
        ws = wb.create_sheet(sheet_title)

    fila = 1

    # Título
    ws.merge_cells(f'A{fila}:{chr(67 + len(años)*2)}{fila}')
    titulo = ws.cell(row=fila, column=1, value=f"COMPARACIÓN POR MES Y AÑO - {toxina}")
    titulo.font = titulo_font
    titulo.fill = titulo_fill
    titulo.alignment = center_alignment
    fila += 2

    # Encabezados de años
    ws.cell(row=fila, column=1, value="MES").font = header_font
    ws.cell(row=fila, column=1).fill = header_fill
    ws.cell(row=fila, column=1).alignment = center_alignment

    col = 2
    for año in años:
        ws.cell(row=fila, column=col, value=f"Cont {año}").font = header_font
        ws.cell(row=fila, column=col).fill = header_fill
        ws.cell(row=fila, column=col).alignment = center_alignment
        ws.cell(row=fila, column=col+1, value=f"% {año}").font = header_font
        ws.cell(row=fila, column=col+1).fill = header_fill
        ws.cell(row=fila, column=col+1).alignment = center_alignment
        col += 2
    fila += 1

    # Datos por mes
    for mes in todos_los_meses:
        ws.cell(row=fila, column=1, value=mes).alignment = left_alignment

        col = 2
        for año in años:
            datos = datos_por_toxina[toxina][año][mes]
            ws.cell(row=fila, column=col, value=datos['contaminados']).alignment = center_alignment
            ws.cell(row=fila, column=col+1, value=datos['porcentaje']).alignment = center_alignment
            col += 2
        fila += 1

    # Fila de totales por año para esta toxina
    ws.cell(row=fila, column=1, value="TOTAL/AÑO").font = Font(bold=True)
    col = 2
    for año in años:
        total_cont = df_anual[(df_anual['Año'] == año) & (df_anual['Micotoxina'] == toxina)]['Contaminados'].values
        total_porc = df_anual[(df_anual['Año'] == año) & (df_anual['Micotoxina'] == toxina)]['% Contaminacion'].values
        if len(total_cont) > 0:
            ws.cell(row=fila, column=col, value=int(total_cont[0])).font = Font(bold=True)
            ws.cell(row=fila, column=col+1, value=total_porc[0]).font = Font(bold=True)
        col += 2

    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 15
    for i in range(len(años)):
        ws.column_dimensions[chr(66 + i*2)].width = 12
        ws.column_dimensions[chr(67 + i*2)].width = 10


# HOJA 4: DETALLE POR MES Y AÑO

ws4 = wb.create_sheet("Detalle_por_Mes_Año")

fila = 1

ws4.merge_cells(f'A{fila}:G{fila}')
titulo = ws4.cell(row=fila, column=1, value="DETALLE POR MES Y AÑO")
titulo.font = titulo_font
titulo.fill = titulo_fill
titulo.alignment = center_alignment
fila += 2

encabezados = ['Año', 'Mes', 'Micotoxina', 'Total Muestras', 'Contaminados', 'No Contaminados', '%']
for col, enc in enumerate(encabezados, 1):
    cell = ws4.cell(row=fila, column=col, value=enc)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_alignment
fila += 1

for año in años:
    for mes in todos_los_meses:
        df_mes_año = df_detalle[(df_detalle['Año'] == año) & (df_detalle['Mes'] == mes)]
        if len(df_mes_año) > 0 and df_mes_año['Total Muestras'].sum() > 0:
            ws4.cell(row=fila, column=1, value=año).font = Font(bold=True)
            ws4.cell(row=fila, column=2, value=mes).font = Font(bold=True)
            ws4.merge_cells(f'C{fila}:G{fila}')
            ws4.cell(row=fila, column=3, value=f"--- DATOS DEL MES {mes} {año} ---").font = Font(bold=True)
            fila += 1

            for _, row in df_mes_año.iterrows():
                if row['Total Muestras'] > 0:
                    ws4.cell(row=fila, column=1, value=row['Año']).alignment = center_alignment
                    ws4.cell(row=fila, column=2, value=row['Mes']).alignment = center_alignment
                    ws4.cell(row=fila, column=3, value=row['Micotoxina']).alignment = left_alignment
                    ws4.cell(row=fila, column=4, value=row['Total Muestras']).alignment = center_alignment
                    ws4.cell(row=fila, column=5, value=row['Contaminados']).alignment = center_alignment
                    ws4.cell(row=fila, column=6, value=row['No Contaminados']).alignment = center_alignment
                    ws4.cell(row=fila, column=7, value=row['% Contaminacion']).alignment = center_alignment
                    fila += 1

            fila += 1

ws4.column_dimensions['A'].width = 10
ws4.column_dimensions['B'].width = 15
ws4.column_dimensions['C'].width = 25
ws4.column_dimensions['D'].width = 15
ws4.column_dimensions['E'].width = 15
ws4.column_dimensions['F'].width = 15
ws4.column_dimensions['G'].width = 10



# HOJA 5: ANÁLISIS DE MÚLTIPLES MICOTOXINAS


ws5 = wb.create_sheet("Multimicotoxinas")

fila = 1

ws5.merge_cells(f'A{fila}:F{fila}')
titulo = ws5.cell(row=fila, column=1, value="ANÁLISIS DE MUESTRAS CON MÚLTIPLES MICOTOXINAS")
titulo.font = titulo_font
titulo.fill = titulo_fill
titulo.alignment = center_alignment
fila += 2

# Resumen general
ws5.cell(row=fila, column=1, value="RESUMEN GENERAL").font = seccion_font
ws5.cell(row=fila, column=1).fill = seccion_fill
fila += 1

ws5.cell(row=fila, column=1, value="Total de muestras analizadas:")
ws5.cell(row=fila, column=2, value=total_muestras)
fila += 1

ws5.cell(row=fila, column=1, value="Muestras con más de 1 micotoxina:")
ws5.cell(row=fila, column=2, value=num_muestras_multitoxina)
fila += 1

ws5.cell(row=fila, column=1, value="Porcentaje de muestras con múltiples micotoxinas:")
ws5.cell(row=fila, column=2, value=f"{round(num_muestras_multitoxina / total_muestras * 100, 2)}%")
fila += 2

# Distribución por número de micotoxinas
ws5.cell(row=fila, column=1, value="DISTRIBUCIÓN POR NÚMERO DE MICOTOXINAS").font = seccion_font
ws5.cell(row=fila, column=1).fill = seccion_fill
fila += 1

ws5.cell(row=fila, column=1, value="N° Micotoxinas")
ws5.cell(row=fila, column=2, value="N° Muestras")
ws5.cell(row=fila, column=3, value="Porcentaje")
for col in range(1, 4):
    ws5.cell(row=fila, column=col).font = header_font
    ws5.cell(row=fila, column=col).fill = header_fill
fila += 1

for n_micotoxinas, count in distribucion.items():
    ws5.cell(row=fila, column=1, value=n_micotoxinas)
    ws5.cell(row=fila, column=2, value=count)
    ws5.cell(row=fila, column=3, value=f"{round(count / total_muestras * 100, 2)}%")
    fila += 1
fila += 2

# Análisis por año
ws5.cell(row=fila, column=1, value="ANÁLISIS POR AÑO").font = seccion_font
ws5.cell(row=fila, column=1).fill = seccion_fill
fila += 1

ws5.cell(row=fila, column=1, value="Año")
ws5.cell(row=fila, column=2, value="Total Muestras")
ws5.cell(row=fila, column=3, value="Muestras con >1 micotoxina")
ws5.cell(row=fila, column=4, value="Porcentaje")
for col in range(1, 5):
    ws5.cell(row=fila, column=col).font = header_font
    ws5.cell(row=fila, column=col).fill = header_fill
fila += 1

for año in años:
    df_año = df[df['AÑO'] == año]
    total_año = len(df_año)
    multitoxina_año = len(df_año[df_año['NUM_MICOTOXINAS'] > 1])
    porcentaje_año = round(multitoxina_año / total_año * 100, 2) if total_año > 0 else 0

    ws5.cell(row=fila, column=1, value=año)
    ws5.cell(row=fila, column=2, value=total_año)
    ws5.cell(row=fila, column=3, value=multitoxina_año)
    ws5.cell(row=fila, column=4, value=f"{porcentaje_año}%")
    fila += 1
fila += 2

# Análisis por mes
ws5.cell(row=fila, column=1, value="ANÁLISIS POR MES").font = seccion_font
ws5.cell(row=fila, column=1).fill = seccion_fill
fila += 1

ws5.cell(row=fila, column=1, value="Mes")
ws5.cell(row=fila, column=2, value="Total Muestras")
ws5.cell(row=fila, column=3, value="Muestras con >1 micotoxina")
ws5.cell(row=fila, column=4, value="Porcentaje")
for col in range(1, 5):
    ws5.cell(row=fila, column=col).font = header_font
    ws5.cell(row=fila, column=col).fill = header_fill
fila += 1

for mes in todos_los_meses:
    df_mes = df[df['MES'] == mes]
    total_mes = len(df_mes)
    if total_mes > 0:
        multitoxina_mes = len(df_mes[df_mes['NUM_MICOTOXINAS'] > 1])
        porcentaje_mes = round(multitoxina_mes / total_mes * 100, 2)

        ws5.cell(row=fila, column=1, value=mes)
        ws5.cell(row=fila, column=2, value=total_mes)
        ws5.cell(row=fila, column=3, value=multitoxina_mes)
        ws5.cell(row=fila, column=4, value=f"{porcentaje_mes}%")
        fila += 1
fila += 2

# Combinaciones más frecuentes
ws5.cell(row=fila, column=1, value="COMBINACIONES MÁS FRECUENTES").font = seccion_font
ws5.cell(row=fila, column=1).fill = seccion_fill
fila += 1

ws5.cell(row=fila, column=1, value="Combinación de micotoxinas")
ws5.cell(row=fila, column=2, value="N° Muestras")
for col in range(1, 3):
    ws5.cell(row=fila, column=col).font = header_font
    ws5.cell(row=fila, column=col).fill = header_fill
fila += 1

for combinacion, count in contador_combinaciones.most_common(20):
    ws5.cell(row=fila, column=1, value=combinacion)
    ws5.cell(row=fila, column=2, value=count)
    fila += 1

# Ajustar ancho de columnas
for col in range(1, 5):
    ws5.column_dimensions[chr(64 + col)].width = 25


# GUARDAR ARCHIVO

wb.save(output_filename)
print(f"Archivo guardado como: {output_filename}")

# MOSTRAR RESULTADOS EN CONSOLA

print("RESULTADOS DEL ANALISIS")
print("CRITERIO: Números y valores con '>' = CONTAMINADO; '<' = NO CONTAMINADO")

print("\n TABLA 1: RESUMEN POR AÑO (MATRIZ)")
print("(Contaminados / % Contaminacion)")


print(f"{'Micotoxina':<20}", end="")
for año in años:
    print(f"  Cont {año}  % {año}  ", end="")
print()


for toxina in toxinas:
    print(f"{toxina[:20]:<20}", end="")
    for año in años:
        valor = df_anual[(df_anual['Año'] == año) & (df_anual['Micotoxina'] == toxina)]['Contaminados'].values
        porcentaje = df_anual[(df_anual['Año'] == año) & (df_anual['Micotoxina'] == toxina)]['% Contaminacion'].values
        if len(valor) > 0:
            print(f"      {int(valor[0]):3}    {porcentaje[0]:5.1f}   ", end="")
        else:
            print(f"      -      -     ", end="")
    print()

print("\n\n TABLA 2: RESUMEN POR MES (MATRIZ)")
print("(Contaminados / % Contaminacion)")


print(f"{'Micotoxina':<20}", end="")
for mes in todos_los_meses:
    print(f"  Cont {mes[:4]}  % {mes[:4]}  ", end="")
print()


for toxina in toxinas:
    print(f"{toxina[:20]:<20}", end="")
    for mes in todos_los_meses:
        valor = df_mensual[(df_mensual['Mes'] == mes) & (df_mensual['Micotoxina'] == toxina)]['Contaminados'].values
        porcentaje = df_mensual[(df_mensual['Mes'] == mes) & (df_mensual['Micotoxina'] == toxina)]['% Contaminacion'].values
        if len(valor) > 0 and valor[0] > 0:
            print(f"      {int(valor[0]):3}    {porcentaje[0]:5.1f}   ", end="")
        else:
            print(f"      0      0.0     ", end="")
    print()

print("\n\nTABLA 3: COMPARACIÓN POR MES Y AÑO (ejemplo primera micotoxina) ")
print("Formato: Mes / Contaminados y % por año")


primera_toxina = toxinas[0]
print(f"\n{primera_toxina}:")
print(f"{'MES':<15}", end="")
for año in años:
    print(f"  Cont {año}  % {año}  ", end="")
print()

for mes in todos_los_meses:
    print(f"{mes:<15}", end="")
    for año in años:
        datos = datos_por_toxina[primera_toxina][año][mes]
        print(f"      {datos['contaminados']:3}    {datos['porcentaje']:5.1f}   ", end="")
    print()

print("\n\nANÁLISIS DE MÚLTIPLES MICOTOXINAS")
print(f"Muestras con más de 1 micotoxina: {num_muestras_multitoxina} de {total_muestras} ({round(num_muestras_multitoxina / total_muestras * 100, 2)}%)")



# DESCARGAR ARCHIVO


print("\nDESCARGANDO ARCHIVO")

files.download(output_filename)



# FINAL DEL PROCESO


print("PROCESO COMPLETADO")

print("\nResumen de resultados:")
print(f"   - Total de muestras analizadas: {len(df)}")
print(f"   - Años analizados: {', '.join(map(str, años))}")
print(f"   - Meses con datos: {', '.join(meses_con_datos) if meses_con_datos else 'Ninguno'}")
print(f"   - Micotoxinas analizadas: {len(columnas_toxinas)}")
print(f"   - Muestras con múltiples micotoxinas: {num_muestras_multitoxina}")
print(f"   - Archivo descargado: {output_filename}")

print("\nESTRUCTURA DEL ARCHIVO EXCEL:")
print("   - Hoja 1: Resumen por Año (matriz años vs toxinas)")
print("   - Hoja 2: Resumen por Mes (matriz 12 meses vs toxinas)")
print("   - Hoja 3: Comparación por Mes y Año (una tabla por micotoxina)")
print("     * Cada micotoxina tiene su propia hoja o sección")
print("     * Filas: meses (ENERO a DICIEMBRE)")
print("     * Columnas: Cont AÑO y % AÑO para cada año")
print("   - Hoja 4: Detalle por Mes y Año (resumen adicional)")
print("   - Hoja 5: Análisis de Múltiples Micotoxinas (NUEVA)")
print("     * Resumen general")
print("     * Distribución por número de micotoxinas")
print("     * Análisis por año")
print("     * Análisis por mes")
print("     * Combinaciones más frecuentes")
